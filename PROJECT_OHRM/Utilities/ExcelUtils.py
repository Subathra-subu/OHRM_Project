import openpyxl
#added excel
def get_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_columns + 1):
            val = sheet.cell(r, c).value
            # Convert None to empty string, and everything else explicitly to string
            row_list.append("" if val is None else str(val).strip())
        final_list.append(row_list)
    return final_list